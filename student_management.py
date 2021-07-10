import os
import tkinter
from tkinter import *
from tkinter.font import *
from tkinter import ttk, messagebox
import sqlite3
import datetime
import webbrowser


def take_input():
    con = sqlite3.connect("Resources/management.db")
    cur1 = con.cursor()
    cur1.execute("""
    CREATE TABLE IF NOT EXISTS students (
        roll_no integer primary key,
        name text,
        marks integer
    )
    """)

    try:
        roll_no = int(entry2.get())
        entry2.delete(0, tkinter.END)
        na_me = str(entry1.get())
        entry1.delete(0, tkinter.END)
        marks = int(entry3.get())
        entry3.delete(0, tkinter.END)

        current_time = datetime.datetime.now()
        with open("Resources/history.txt", "a") as hist_file:
            hist_file.write(f'Data entered at {current_time.strftime("%d-%m-%Y %H:%M:%S")}:\n')
            hist_file.write(f'Roll_No:{roll_no} Name:"{na_me}" Marks:{marks}\n\n')

        cur2 = con.cursor()
        query = f"""INSERT INTO students VALUES ({roll_no}, "{na_me}", {marks})"""
        try:
            cur2.execute(query)
            con.commit()
            con.close()
            messagebox.showinfo("Success", "Data Recorded Successfully")
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", f"The Roll Number {roll_no} is Already Assigned!")
        except sqlite3.OperationalError:
            messagebox.showerror("Error", "Enter a Proper Name")
    except ValueError:
        entry1.delete(0, tkinter.END)
        entry2.delete(0, tkinter.END)
        entry3.delete(0, tkinter.END)
        messagebox.showerror("Error", "Please Enter Correct Data!")


def display_list():
    def add_record():
        root.state("normal")
        new_win.destroy()

    def remove_record():
        try:
            var = tree.selection()[0]
            tree.delete(var)
            con3 = sqlite3.connect("Resources/management.db")
            t_cur = con3.cursor()
            t_cur.execute(f"DELETE FROM students WHERE roll_no={int(stu[int(var)])};")
            con3.commit()
            con3.close()
        except IndexError:
            messagebox.showerror("Error", f"Please Select Any Record!!")

    def open_in_excel():
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
            import os
            wb = Workbook()
            ws = wb.active

            ws.merge_cells('A1:C1')
            t_cell = ws.cell(row=1, column=1)
            t_cell.value = "Student Data"
            t_cell.alignment = Alignment(horizontal='center', vertical='center')
            t_cell.font = Font(name='Century Gothic',
                               size=14,
                               bold=True,
                               color='651778')

            exl_con = sqlite3.connect("Resources/management.db")
            exl_cur = exl_con.cursor()
            exl_cur.execute("SELECT * FROM Students")

            for exl_i in exl_cur:
                ws.append([exl_i[0], exl_i[1], exl_i[2]])

            exl_con.commit()
            exl_con.close()
            wb.save("Excel_Files/exl.xlsx")

            os.chdir("Excel_Files")
            os.system('exl.xlsx')
            os.chdir("..")
        except ModuleNotFoundError:
            messagebox.showwarning("Requirements",
                                   "To enable this feature run 'Exl_Requirements.bat'"
                                   )

    def help_stm():
        os.system("README.md")

    def rate_us():
        webbrowser.open("https://forms.gle/vyxnnkRb7JTMPeSN9", new=1)

    # Minimizing the main window
    # And setting up a new window
    root.state("iconic")
    new_win = Tk()
    width_n, height_n = 800, 320
    screen_width_n = new_win.winfo_screenwidth()
    screen_height_n = new_win.winfo_screenheight()
    x_coordinate_n = (screen_width_n / 2) - (width_n / 2)
    y_coordinate_n = (screen_height_n / 2) - (height_n / 2)
    # Making the window appear at the center of the Screen
    new_win.geometry(f"{width_n}x{height_n}+{int(x_coordinate_n)}+{int(y_coordinate_n)}")
    new_win.title("Student Data")
    try:
        new_win.wm_iconbitmap("Resources/stud_img.ico")
    except Exception:
        pass

    # -- Menu --
    menu = Menu(new_win)

    m1 = Menu(menu, tearoff=0)
    m1.add_command(label="Open in Excel", command=open_in_excel)
    # m1.add_separator()
    # m1.add_command(label="Exit", command=exit)

    menu.add_command(label="🠔", command=add_record)
    menu.add_cascade(label="Options", menu=m1)
    menu.add_command(label="Help", command=help_stm)
    menu.add_command(label="Rate Us", command=rate_us)
    new_win.configure(menu=menu)

    # -- Heading --
    app_label = Label(new_win, text="Student Management System", fg="VioletRed1", width=40)
    app_label.config(font=("Times New Roman", 30, BOLD), bg="antique white")
    app_label.pack()

    # Table Styles
    style = ttk.Style(new_win)
    style.configure('Treeview',
                    background="LightBlue1",
                    rowheight=30
                    )
    style.configure("Treeview", font=("Century", 12))
    style.map('Treeview',
              background=[("selected", "green")]
              )
    # Frame 1
    frame_1 = Frame(new_win, bg="antique white")
    frame_1.pack()
    # Table Configuration
    tree = ttk.Treeview(frame_1)
    tree["columns"] = ("one", "two", "three")
    tree.heading("one", text="Roll Number")
    tree.heading("two", text="Name")
    tree.heading("three", text="Marks")
    # Getting Student Data From The Database
    con1 = sqlite3.connect("Resources/management.db")
    cur_n = con1.cursor()
    cur_n.execute("SELECT * FROM students")
    i = 0
    # Customizing Columns
    tree.column("#0", width=0, stretch=NO)
    tree.column("one", anchor=CENTER, width=80)
    tree.column("two", anchor=CENTER, width=300)
    tree.column("three", anchor=CENTER, width=120)
    stu = []
    # Inserting Data Into The Table
    for rows in cur_n:
        tree.insert(parent='', index='end', iid=i, text="", values=(rows[0], rows[1], rows[2]))
        stu.append(rows[0])
        i += 1
    tree.config(height=5)
    # Scrollbar For Table
    scr = ttk.Scrollbar(frame_1, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scr.set)
    tree.pack(pady=10, side=LEFT)
    scr.pack(side=RIGHT, fill=Y)

    # Frame 2
    frame_2 = Frame(new_win, bg="antique white")
    frame_2.pack()
    btn_font = Font(size=12, family='Century', slant=ITALIC)
    add_button = Button(frame_2, text="Add Record", bg="tan1", font=btn_font, command=add_record)
    add_button.grid(row=1, column=1, padx=20, pady=15)
    del_button = Button(frame_2, text="Delete Record", bg="tan1", font=btn_font, command=remove_record)
    del_button.grid(row=1, column=3, padx=20, pady=15)

    new_win.configure(bg="antique white")
    new_win.mainloop()


def show_history():
    def gnt_text():
        import os
        os.chdir('Resources')
        os.system('history.txt')
        os.chdir('..')

    def del_hist():
        with open("Resources/history.txt", "w") as file:
            file.write("")
        hist_win.destroy()
        root.state("normal")
        messagebox.showinfo("Success", "History Records are cleared Successfully.")

    def h_go_back():
        hist_win.destroy()
        root.state("normal")

    root.state("iconic")
    hist_win = Tk()
    hist_win.title("History")
    h_width, h_height = 600, 340
    h_screen_width = hist_win.winfo_screenwidth()
    h_screen_height = root.winfo_screenheight()
    x_coordinate_h = (h_screen_width / 2) - (h_width / 2)
    y_coordinate_h = (h_screen_height / 2) - (h_height / 2)
    hist_win.geometry(f"{h_width}x{h_height}+{int(x_coordinate_h)}+{int(y_coordinate_h)}")
    hist_win.maxsize(width="600", height="340")
    try:
        hist_win.wm_iconbitmap("Resources/stud_img.ico")
    except Exception:
        pass

    # --Menu--
    h_menu = Menu(hist_win)
    h_menu.add_command(label="Back", command=h_go_back)
    hist_win.configure(menu=h_menu)

    read_hist = open("Resources/history.txt", "r")
    hist = read_hist.readlines()

    h_frame = Frame(hist_win, bg="SeaGreen1")
    h_frame.pack()
    scrollbar = Scrollbar(h_frame)
    scrollbar.pack(side=RIGHT, fill=Y)
    lst_font = Font(size=12, family="Times New Roman")

    my_list = Listbox(h_frame, yscrollcommand=scrollbar.set, height=13)
    if len(hist) != 0:
        for i in hist:
            my_list.insert(END, i)
    else:
        my_list.insert(END, "---No History Found!!---")
    my_list.configure(width=100, font=lst_font, bg="khaki1")
    my_list.pack(pady=20)
    scrollbar.config(command=my_list.yview)
    read_hist.close()

    h_frame2 = Frame(hist_win, bg="SeaGreen1")
    h_frame2.pack()
    txt_btn = Button(h_frame2, text="Generate Text File", font=("Century Gothic", 12), bg='OliveDrab2',
                     command=gnt_text)
    txt_btn.pack(side=LEFT, padx=30)
    del_hist_btn = Button(h_frame2, text="Clear History", font=("Century Gothic", 12), bg='OliveDrab2',
                          command=del_hist)
    del_hist_btn.pack(side=RIGHT, padx=30)

    hist_win.configure(bg="SeaGreen1")
    hist_win.mainloop()


root = Tk()
root.title("Student Management System")
width, height = 800, 390
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x_coordinate = (screen_width / 2) - (width / 2)
y_coordinate = (screen_height / 2) - (height / 2)
root.geometry(f"{width}x{height}+{int(x_coordinate)}+{int(y_coordinate)}")
root.minsize(width=width, height=height)
root.maxsize(width=width, height=height)
try:
    root.wm_iconbitmap("Resources/stud_img.ico")
except Exception:
    pass

img = PhotoImage(file='Resources/stud_img.png')
test_l = Label(root, image=img)
test_l.place(x=0, y=0)

# -- Menu --
m_menu = Menu(root)
opt_m = Menu(m_menu, tearoff=0)
opt_m.add_command(label="History", command=show_history)
# opt_m.add_separator()
# opt_m.add_command(label="Exit", command=exit)

m_menu.add_cascade(label="Options", menu=opt_m)
root.configure(menu=m_menu)

heading_style = Font(family="Times New Roman", size=19, weight=BOLD)
heading = Label(root, text="Welcome to the Student Management System!!", font=heading_style, fg='blue2')
heading.pack(pady=12, padx=12)
fr1 = Frame(root, borderwidth=3, bg='salmon1', relief=SUNKEN)
fr1.pack(ipady=10, ipadx=15, pady=15)

font_style = Font(size=20, family='Monotype Corsiva', weight=BOLD)
entry_style = Font(size=16, family='Century')
btn_style = Font(size=12, family='Century', slant=ITALIC)

l1 = Label(fr1, text="Name:", bg='salmon1', font=font_style)
l1.grid(row=1, column=1, padx=12, pady=12)
entry1 = Entry(fr1, width='40', font=entry_style)
entry1.grid(row=1, column=2, pady=12)

l2 = Label(fr1, text="Roll Number:", bg='salmon1', font=font_style)
l2.grid(row=2, column=1, padx=12, pady=12)
entry2 = Entry(fr1, width='40', font=entry_style)
entry2.grid(row=2, column=2, pady=12)

l3 = Label(fr1, text="Marks:", bg='salmon1', font=font_style)
l3.grid(row=3, column=1, padx=12, pady=12)
entry3 = Entry(fr1, width='40', font=entry_style)
entry3.grid(row=3, column=2, pady=12)

fr2 = Frame(root, bg='antique white')
fr2.pack(ipadx=0, ipady=0)
btn1 = Button(fr2, text="Take Input", font=btn_style, bg='DarkOliveGreen1', command=take_input)
btn1.grid(row=1, column=1, padx='50', pady='15')
btn2 = Button(fr2, text="Display List", font=btn_style, bg='DarkOliveGreen1', command=display_list)
btn2.grid(row=1, column=2, padx='50', pady='15')

root.mainloop()
